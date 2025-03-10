import os
import json
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from math import sqrt



def load_cases_numbers_to_excel(start_cell, num_lc, repetitions, space, gap = 1, mode = "default" ,sheet_to_use = "Json Input"):
    sheet = workbook[sheet_to_use]
    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    for i in range(repetitions):

        # Itera su ogni numero da 1 a num_values
        for value in range(1, num_lc + 1):
            # Calcola la colonna corrente
            current_column = start_column + ((i * space) + ((value - 1) * gap)) #+ (gap * value)
            #print(current_column)

            # Scrivi il valore nella cella corrispondente
            sheet.cell(row=start_row, column=current_column, value=value)

            if mode == "mt":
                sheet.cell(row=start_row, column=current_column+1, value=value)


def piles_numbers_to_excel(start_cell, num_piles, repetitions, space, gap = 1, sheet_to_use = "Json Input"):
    sheet = workbook[sheet_to_use]
    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    for i in range(repetitions):
        current_column = start_column + ( i * space)
        #row == start_row
        # Itera su ogni numero da 1 a num_values
        for value in range(1, num_piles + 1):
            # Calcola la colonna corrente
            current_row = start_row + (value - 1) * gap #+ (gap * value)
            #print(current_column)
            #row += 1
            # Scrivi il valore nella cella corrispondente
            sheet.cell(row=current_row, column=current_column, value=value)

def write_pairs_to_excel(start_cell, data_list, sheet_to_use = "Json Input"):
    # Carica il file Excel esistente
    
    sheet = workbook[sheet_to_use]

    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    # Itera su ogni elemento della lista
    for index, value in enumerate(data_list, start=1):
        # Calcola la colonna corrente
        current_row = start_row + (index - 1)
        

        # Scrivi il numero crescente nella cella corrispondente
        sheet.cell(row=current_row, column=start_column, value=index)

        # Scrivi il valore dalla lista nella cella successiva
        sheet.cell(row=current_row, column=start_column + 1, value=value)


def write_titles_to_excel(start_cell, num_rows, json_list, key_name, repetitions, space, gap, repeat = False, sheet_to_use = "Json Input"):
    # Carica il file Excel esistente e specifica il foglio di lavoro
    sheet = workbook[sheet_to_use]

    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row = sheet[start_cell].row
    start_column = sheet[start_cell].column

    # Inizializza la variabile per tenere traccia del numero di valori scritti
    values_written = 0

    # Inizializza la variabile per tenere traccia della colonna corrente
    #current_column = start_column

    for i in range(repetitions):
        current_column = start_column + (space * i) 

        for i in json_list:
            #print(i)
            column_to_use = current_column + (values_written * gap)
            val_to_use = i[key_name]



            # Scrivi il valore nella colonna corrente sulla riga corrente
            sheet.cell(row=start_row , column=column_to_use, value=val_to_use)

            if repeat == True:
                sheet.cell(row=start_row , column=column_to_use+1, value=val_to_use)

            #sheet.cell(row=start_row + values_written, column=current_column+1, value=val_to_use)

            # Incrementa il contatore dei valori scritti
            values_written += 1

            # Se il numero di valori scritti è uguale al numero di pile, passa alla colonna successiva
            if values_written == num_rows:
                values_written = 0  # Resetta il contatore
                current_column += 1    # Passa alla colonna successiva


def write_values_to_excel(start_cell, num_rows, json_list, key_name, sheet_to_use = "Json Input", iterate= True, mode = "deafult", axial= 0):
    # Carica il file Excel esistente e specifica il foglio di lavoro
    sheet = workbook[sheet_to_use]

    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row = sheet[start_cell].row
    start_column = sheet[start_cell].column

    # Inizializza la variabile per tenere traccia del numero di valori scritti
    values_written = 0

    # Inizializza la variabile per tenere traccia della colonna corrente
    current_column = start_column

    #MN TAB
    mn_row = sheet["CX44"].row
    mn_column= sheet["CX44"].column
    mn_current_col = mn_column

    #TN TAB
    tn_row= sheet["DX44"].row
    tn_column= sheet["DX44"].column
    tn_current_col = tn_column
    axialload_list = []


    if iterate == True: 

        # Itera su ciascun elemento di json list
        for i in json_list:
            # Itera su ciascun dizionario all'interno di jsonlist
            for j in i:
                # Accedi al valore specificato dalla key_name
                #print(j)
                val_to_use = j[key_name]

                #name_of_list.append(val_to_use)

                # Scrivi il valore nella colonna corrente sulla riga corrente
                sheet.cell(row=start_row + values_written, column=current_column, value=val_to_use)
                
                if mode == "mn and tn":
                    sheet.cell(row=mn_row + values_written, column=mn_current_col, value=val_to_use)
                    

                    sheet.cell(row=tn_row + values_written, column=tn_current_col, value=val_to_use)
                    

                # Incrementa il contatore dei valori scritti
                values_written += 1

                # Se il numero di valori scritti è uguale al numero di pile, passa alla colonna successiva
                if values_written == num_rows:
                    values_written = 0  # Resetta il contatore
                    current_column += 1    # Passa alla colonna successiva
                    tn_current_col += 2
                    mn_current_col += 2
                #return name_of_list
    #print(axialload_list)
    if axial == 1:
        for i in json_list:
            # Itera su ciascun dizionario all'interno di jsonlist
            for j in i:
                # Accedi al valore specificato dalla key_name
                #print(j)
                val_to_use = j[key_name]
                axialload_list.append(val_to_use)



        
        #print(axialload_list)
        massimo = int(max(axialload_list))
        centinaia_superiore = ((massimo // 100) + 1) * 100

        limite_taglio_row, limite_taglio_col = sheet["EV44"].row, sheet["EV44"].column

        

        nuova_lista = list(range(100, centinaia_superiore + 1,100))
        #print(massimo)
        #print(nuova_lista)
        #val_written = 0
        for ind, i in enumerate(nuova_lista, start= 1):
            #print(ind)
            curr_row = limite_taglio_row + (ind - 1)
            #print(curr_row)
            
            sheet.cell(row=curr_row, column=limite_taglio_col, value=i*0.2)
            sheet.cell(row=curr_row, column=limite_taglio_col+1, value=i)
            #val_written += 1

        

def sqroot_of_lists(first_list, key_name1, second_list, key_name2, start_cell, num_rows, mode = "default", sheet_to_use = "Json Input"):
    first_values = []
    second_values = []
    square_roots = []

    for i in first_list:
        # Itera su ciascun dizionario all'interno di jsonlist
        for j in i:
            # Accedi al valore specificato dalla key_name
            #print(j)
            first_values.append(j[key_name1])

    for i in second_list:
        # Itera su ciascun dizionario all'interno di jsonlist
        for j in i:
            # Accedi al valore specificato dalla key_name
            #print(j)
            second_values.append(j[key_name2])

    for i, element in enumerate(first_values):
        a = (first_values[i]*first_values[i] + second_values[i]*second_values[i])**0.5
        square_roots.append(a) #(first_values[i]*first_values[i] + second_values[i]*second_values[i])**0.5) #(first_values[i]**2 + second_values[i]**2)**0.5)
        
    
    
    sheet = workbook[sheet_to_use]

    
    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row = sheet[start_cell].row
    start_column = sheet[start_cell].column

    # Inizializza la variabile per tenere traccia del numero di valori scritti
    values_written = 0

    # Inizializza la variabile per tenere traccia della colonna corrente
    current_column = start_column

    # Itera su ciascun elemento di json list
    for i in square_roots:
        # Itera su ciascun dizionario all'interno di jsonlist
        #for j in i:
        # Accedi al valore specificato dalla key_name
        #print(j)
        val_to_use = i

        #name_of_list.append(val_to_use)
        #return name_of_list
        
        # Scrivi il valore nella colonna corrente sulla riga corrente
        sheet.cell(row=start_row + values_written, column=current_column, value=val_to_use)

        # Incrementa il contatore dei valori scritti
        values_written += 1

        # Se il numero di valori scritti è uguale al numero di pile, passa alla colonna successiva
        if values_written == num_rows:
            values_written = 0  # Resetta il contatore
            current_column += 1    # Passa alla colonna successiva

    if mode == "mn":
        start_row = sheet["CW44"].row
        start_column = sheet["CW44"].column
        current_column = start_column
        values_written = 0

        for i, numb in enumerate(square_roots):
            val_to_use = numb
            
            sheet.cell(row=start_row + values_written, column=current_column, value=val_to_use )
            values_written += 1
            if values_written == num_rows:
                values_written = 0  # Resetta il contatore
                current_column += 2

    elif mode == "tn":
        start_row = sheet["DW44"].row
        start_column = sheet["DW44"].column
        current_column = start_column
        values_written = 0

        for i, numb in enumerate(square_roots):
            val_to_use = numb
            
            sheet.cell(row=start_row + values_written, column=current_column, value=val_to_use )
            values_written += 1
            if values_written == num_rows:
                values_written = 0  # Resetta il contatore
                current_column += 2

def pile_main_info(list, start_cell, sheet_to_use = "Json Input"):
    

    sheet = workbook[sheet_to_use]

    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    for ind, i in enumerate(list, start=1):
        current_row = start_row + (ind-1)
        x = i["x coord"]["Value"]
        y = i["y coord"]["Value"]
        shaft_diam = i["Shaft diameter"]["Value"]
        base_diam = i["Base diameter"]["Value"]
        emb_length = i["Embedded length"]["Value"]

        sheet.cell(row=current_row, column= start_column, value=ind)
        sheet.cell(row=current_row, column= start_column+1, value=shaft_diam)
        sheet.cell(row=current_row, column= start_column+2, value=base_diam)
        sheet.cell(row=current_row, column= start_column+3, value=x)
        sheet.cell(row=current_row, column= start_column+4, value=y)
        sheet.cell(row=current_row, column= start_column+5, value=emb_length)


def qult_comp(list, start_cell, sheet_to_use = "Json Input"):
    sheet = workbook[sheet_to_use]
    axial_cap_list=[]

    qult_comp_list = []
    qult_comp_ind_list =[]
    length_list =[]
    length_list_uniq = []
    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    tens_comp_ratio= data["Inputs"]["Pile data"]["Tensile / compressive axial capacity ratio"]["Value"]
    pile_head_responses = data['Outputs']['Pile head response']

    for i in list:
        length = i["Embedded length"]["Value"]
        length_list.append(length)

    for ind, i in enumerate(list, start=1):
        
        #print(i)
        axial_cap_lim = i["Axial capacity limits"][0]["Axial capacity limit"]
        axial_cap_list.append(axial_cap_lim) #

    #print(axial_cap_list)
    #print(axial_cap_lim)

    dizionario_valori_unici = {}

    for ind, val in enumerate(axial_cap_list):
        if val not in dizionario_valori_unici:
            dizionario_valori_unici[val] = [ind]
        else:
            dizionario_valori_unici[val].append(ind)

    for i in dizionario_valori_unici.values():
        
        qult_comp_ind_list.append(i[0])
        
        #print(i[0])


    #print(dizionario_valori_unici)
    for i in dizionario_valori_unici.keys():
        qult_comp_list.append(i)
        #print(i)

    #print("compressione:" + str(qult_comp_list))
    #print("Ind:" + str(qult_comp_ind_list))
    #print("ratio:" + str(tens_comp_ratio))

    for i in qult_comp_ind_list:
        uniq_length = length_list[i]
        length_list_uniq.append(uniq_length)

    #print("Leng uniq:" + str(length_list_uniq))

    #??? ESTRARRE VALORI UNICI E IL LORO INDCIE PER OTTENERE LUNGHEZZA PALO
    
    for ind, i in enumerate(qult_comp_ind_list, start=1): #dizionario_valori_unici:
        current_row = start_row + (ind-1)
        sheet.cell(row=current_row, column= start_column, value=length_list_uniq[ind-1])
        sheet.cell(row=current_row, column= start_column+1, value=qult_comp_list[ind-1])
        sheet.cell(row=current_row, column= start_column+2, value=((qult_comp_list[ind-1])*(tens_comp_ratio)))

    ratio_row, ratio_col = sheet["S12"].row, sheet["S12"].column
    sheet.cell(row=ratio_row, column=ratio_col, value=tens_comp_ratio)
    
    
def deflect_baricent(start_cell, data_list, total_def, sheet_to_use = "Json Input"):
    # Carica il file Excel esistente
    vert_def_list = []
    horiz_defx_list = []
    rotxz_list = []
    horiz_defy_list = []
    rotyz_list = []
    rotxy_list = []

    sheet = workbook[sheet_to_use]

    # Converte la cella di partenza nella riga e colonna corrispondenti
    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    for i in total_def:
        vert_def = i["Vertical deflection"]
        vert_def_list.append(vert_def)

        horiz_defx = i["Horizontal deflection x"]
        horiz_defx_list.append(horiz_defx)

        rotxz = i["Rotation x to z"]
        rotxz_list.append(rotxz)

        horiz_defy = i["Horizontal deflection y"]
        horiz_defy_list.append(horiz_defy)

        rotyz = i["Rotation y to z"]
        rotyz_list.append(rotyz)

        rotxy = i["Rotation x to y"]
        rotxy_list.append(rotxy)

    #print(vert_def_list)



    for index, value in enumerate(data_list, start=1):
        # Calcola la colonna corrente
        current_row = start_row + (index - 1)
        

        # Scrivi il numero crescente nella cella corrispondente
        sheet.cell(row=current_row, column=start_column, value=index)
        sheet.cell(row=current_row, column=start_column+1, value=vert_def_list[index-1])
        sheet.cell(row=current_row, column=start_column+2, value=horiz_defx_list[index-1])
        sheet.cell(row=current_row, column=start_column+3, value=rotxz_list[index-1])
        sheet.cell(row=current_row, column=start_column+4, value=horiz_defy_list[index-1])
        sheet.cell(row=current_row, column=start_column+5, value=rotyz_list[index-1])
        sheet.cell(row=current_row, column=start_column+6, value=rotxy_list[index-1])


def mn_armature(start_cell, input_path, sheet_to_use = "Json Input"):
    workbook_input = load_workbook(input_path)
    sheet_input = workbook_input.worksheets[5]

    row = sheet_input.max_row
    col = sheet_input.max_column

    sheet = workbook[sheet_to_use]

    start_row, start_column = sheet[start_cell].row, sheet[start_cell].column

    for i in range(1, row +1):
        for j in range(1, col+1):
            c = sheet_input.cell(row= i, column=j)

            sheet.cell(row=start_row+i, column=start_column+j, value=c.value)

    workbook_input.close()






# Ottieni il percorso della directory corrente (dove si trova il codice)
current_directory = os.path.dirname(sys.executable) #os.path.dirname(os.path.abspath(__file__))
#print("curr: " + str(current_directory))
#print(current_directory)

#OTTENERE PATH DI EXCEL DI INPUT
inputxlsx_path = current_directory[:-13]
excelinput_name = []
for i in os.listdir(inputxlsx_path):
    if i.endswith("_piglet_input.xlsx"):
        excelinput_name.append(i)
input_fullpath = os.path.join(inputxlsx_path, excelinput_name[0])
#print(input_fullpath)



# Definisci i nomi delle sottocartelle
diagrams_folder = os.path.join(current_directory, "DIAGRAMS")
json_folder = os.path.join(current_directory, "JSON")

diagrams_names = []
json_names = []
corresponding_files = []

try:
    # Ottieni i nomi dei file nella cartella DIAGRAMS
    for diagname in os.listdir(diagrams_folder):
        diagrams_names.append(diagname)
        #print(diagname)

    # Ottieni i nomi dei file nella cartella JSON
    for jsonname in os.listdir(json_folder):
        json_names.append(jsonname)
        #print(jsonname)

    # Cerca i file corrispondenti
    for json_name in json_names:
        for diag_name in diagrams_names:
            if diag_name.startswith(json_name[:-5]):  # Confronto usando startswith
                json_path = os.path.join(json_folder, json_name)
                diag_path = os.path.join(diagrams_folder, diag_name)
                corresponding_files.append([json_path, diag_path])
                break  # Esci dal loop interno se trovi una corrispondenza

except Exception as ex:
    print(f"Si è verificato un errore: {ex}")


iterat_numb = (str(len(corresponding_files)))

print("There are " + str(iterat_numb) + " iterations associated with the given excel input.")



    
for file_ind, files in enumerate(corresponding_files):
    # Fornisci la directory del file JSON
    json_directory = files[0]
    
    

    # Fornisci la directory del file Excel esistente
    existing_excel_file = files[1]

    try:
        # Carica il contenuto del file JSON
        with open(json_directory, 'r') as json_file:
            data = json.load(json_file)

        # Accedi ai valori 
        pile_head_responses = data['Outputs']['Pile head response']

        loadcaseid = data.get("Inputs", {}).get("Load cases", []) #['Inputs']['Load cases']
        load_cases_id = []
        for i in loadcaseid:
            id = i["Load case ID"]
            load_cases_id.append(id)
            
        #print(load_cases_id)  
        
        #print(loadcaseid)
            
        #INPUT DIAM, X,Y,LENGTH
        pile_info = data.get("Inputs", {}).get("Group data", []) #['Inputs']['Load cases']

        
        # Accedi al numero di load cases e piles da Inputs
        num_piles = data['Inputs']['No. piles']
        num_load = data['Inputs']['No. load cases']
        pile_cap_response = data["Outputs"]["Pile cap response"]

        if existing_excel_file.endswith(".xlsx"):
            corr_existing_excel_file = existing_excel_file
        else:
            corr_existing_excel_file = existing_excel_file + ".xlsx"


        print(corr_existing_excel_file)

        # Carica il file Excel esistente 
        workbook = load_workbook(corr_existing_excel_file) #, keep_vba= True)

        
        


        


        
        
        #numero e nome load cases in tabella in alto
        write_pairs_to_excel(start_cell= "F5", data_list=load_cases_id)

        #ARMATURE MN
        mn_armature(start_cell="I1", input_path=input_fullpath, sheet_to_use="MN")
        
        #SHAFT DIAM, BASE DIAM, X, Y, EMB LENGTH
        pile_main_info(pile_info, "K5")

        #QULT COMP E CULT TENS
        try:
            qult_comp(list= pile_info, start_cell= "R5")
        except:
            pass
        
        #numeri dei load cases nelle tabelle
        load_cases_numbers_to_excel(start_cell= "C43", num_lc = num_load, repetitions = 7, space = 14)
        load_cases_numbers_to_excel(start_cell= "CW43", num_lc = num_load, repetitions = 2, space= 26, gap=2, mode= "mt" )


        #numeri piles nelle tabelle
        piles_numbers_to_excel(start_cell= "B44", num_piles = num_piles, repetitions = 7, space = 14)
        piles_numbers_to_excel(start_cell= "CV44", num_piles = num_piles, repetitions = 2, space = 26)

        #total deflection
        deflect_baricent(start_cell="HX44", data_list= load_cases_id, total_def=pile_cap_response, sheet_to_use = "Json Input")


        #title ax load
        write_titles_to_excel(start_cell= "C42", num_rows= 1, json_list= loadcaseid, key_name= "Load case ID", repetitions= 7, space= 14, gap = 1, sheet_to_use= "Json Input")
        write_titles_to_excel(start_cell= "CW41", num_rows= num_load, json_list= loadcaseid, key_name= "Load case ID", repetitions= 2, space= 26, gap = 2, repeat = True, sheet_to_use= "Json Input")        
        
        #AXIAL LOAD
        write_values_to_excel(start_cell= "C44", num_rows= num_piles, json_list= pile_head_responses, key_name= "Axial load", sheet_to_use= "Json Input", mode = "mn and tn", axial= 1)
        
        #SHEAR LOAD IN X DIRECTION
        write_values_to_excel(start_cell= "Q44", num_rows= num_piles, json_list= pile_head_responses, key_name= "Horizontal load x", sheet_to_use= "Json Input")

        #SHEAR LOAD IN Y DIRECTION
        write_values_to_excel(start_cell= "AE44", num_rows= num_piles, json_list= pile_head_responses, key_name= "Horizontal load y", sheet_to_use= "Json Input")

        #TOTAL SHEAR LOAD
        sqroot_of_lists(first_list = pile_head_responses, key_name1="Horizontal load x", second_list=pile_head_responses, key_name2="Horizontal load y", start_cell= "AS44", num_rows=num_piles, mode = "tn")

        #MOMENT X to Z
        write_values_to_excel(start_cell= "BG44", num_rows= num_piles, json_list= pile_head_responses, key_name= "Moment x to z", sheet_to_use= "Json Input")

        #MOMENT Y to Z
        write_values_to_excel(start_cell= "BU44", num_rows= num_piles, json_list= pile_head_responses, key_name= "Moment y to z", sheet_to_use= "Json Input")

        #RESULTANT MOMENT FOR MN PLOTS
        sqroot_of_lists(first_list = pile_head_responses, key_name1="Moment x to z", second_list=pile_head_responses, key_name2="Moment y to z", start_cell= "CI44", num_rows=num_piles, mode = "mn")


        
        # Salva le modifiche nel file Excel

        workbook.save(corr_existing_excel_file)


        
        
        

    except FileNotFoundError:
        print(f"File non trovato: {json_directory}")
    except KeyError as e:
        print(f"La chiave '{e.args[0]}' non è presente nel file JSON.")
    except json.JSONDecodeError as e:
        print(f"Errore nel decodificare il file JSON: {e}")
    except Exception as ex:
        print(f"Si è verificato un errore: {ex}")


    print( str(file_ind + 1) + " out of " + str(iterat_numb) + " iterations processed.")

print("The diagrams are ready for consultation here: " + str(diagrams_folder))